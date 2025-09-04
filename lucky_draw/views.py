from django.http import HttpResponse, JsonResponse, FileResponse
from io import BytesIO, StringIO
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.db import transaction
import pandas as pd
import random
import json
from .models import Staff, CheckIn, Winner, EventSettings
from .forms import StaffUploadForm, QRCodeScanForm, EventSettingsForm, StaffManualForm
from django.utils import timezone
from django.conf import settings
from datetime import date

# SVG preview endpoint for QR code
@login_required
def preview_qr_svg(request, staff_id):
    staff = get_object_or_404(Staff, id=staff_id)
    import qrcode
    try:
        from qrcode.image.svg import SvgImage
    except ImportError:
        return HttpResponse("SVG QR code generation requires qrcode >=6.1. Please upgrade the qrcode package.", status=500)
    qr_data = f"{staff.id}:{staff.day_1}"
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    svg_img = qr.make_image(image_factory=SvgImage)
    svg_bytes = BytesIO()
    svg_img.save(svg_bytes)
    svg_bytes.seek(0)
    label = f"{staff.name} | {staff.department} | Day {staff.day_1}"
    svg_content = svg_bytes.getvalue().decode('utf-8')
    import re
    m = re.search(r'viewBox="(\d+) (\d+) (\d+) (\d+)"', svg_content)
    if m:
        x, y, w, h = map(int, m.groups())
        label_y = h + 30
        label_x = w // 2
        label_svg = f'<text x="{label_x}" y="{label_y}" text-anchor="middle" font-size="22" fill="#222">{label}</text>'
        svg_content = svg_content.replace('</svg>', label_svg + '</svg>')
    return HttpResponse(svg_content, content_type='image/svg+xml')

import os
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"
os.environ["OMP_NUM_THREADS"] = "1"



@login_required
def download_qr_with_label(request, staff_id):
    staff = get_object_or_404(Staff, id=staff_id)
    import qrcode
    from io import BytesIO
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return HttpResponse("JPG QR code generation requires Pillow. Please install pillow.", status=500)

    # Get event settings
    event_settings = EventSettings.get_solo()
    if not event_settings:
        return HttpResponse("Event settings not found. Please set event dates first.", status=400)

    # Get the date for the specific day
    if staff.day_1 == 1:
        event_date = event_settings.day1_date
        template_path = 'static/templates/day-1-lucky-draw.png'
        qr_coords = (504, 1569, 1083, 2151)
        date_coords = (459, 1331, 1125, 1422)
    elif staff.day_1 == 2:
        event_date = event_settings.day2_date
        template_path = 'static/templates/day-2-lucky-draw.png'
        qr_coords = (502, 1570, 1083, 2152)
        date_coords = (451, 1331, 1106, 1422)
    else:
        return HttpResponse("Invalid day value.", status=400)

    if not event_date:
        return HttpResponse(f"Date for Day {staff.day_1} not set.", status=400)

    # Load the template image
    template_full_path = os.path.join(settings.BASE_DIR, template_path)
    if not os.path.exists(template_full_path):
        return HttpResponse("Template image not found.", status=404)

    template_img = Image.open(template_full_path)
    draw = ImageDraw.Draw(template_img)

    # Generate QR code
    qr_data = f"{staff.id}:{staff.day_1}"
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert('RGB')

    # Resize QR code to fit the coordinates
    qr_width = qr_coords[2] - qr_coords[0]
    qr_height = qr_coords[3] - qr_coords[1]
    qr_img = qr_img.resize((qr_width, qr_height), Image.Resampling.LANCZOS)

    # Paste QR code onto template
    template_img.paste(qr_img, (qr_coords[0], qr_coords[1]))

    # Add date text
    try:
        # Use Poppins ExtraBold font for maximum thickness
        font_path = os.path.join(settings.BASE_DIR, 'static/fonts/Poppins-ExtraBold.ttf')
        if os.path.exists(font_path):
            # Convert 63.7pt to pixels (approximately 84 pixels at 72 DPI)
            font_size = int(63.7 * 1.33)  # Rough conversion from pt to px
            font = ImageFont.truetype(font_path, font_size)
        else:
            # Try Poppins Bold as fallback
            font_path_bold = os.path.join(settings.BASE_DIR, 'static/fonts/Poppins-Bold.ttf')
            if os.path.exists(font_path_bold):
                font_size = int(63.7 * 1.33)
                font = ImageFont.truetype(font_path_bold, font_size)
            else:
                # Try regular Poppins as fallback
                font_path_regular = os.path.join(settings.BASE_DIR, 'static/fonts/Poppins-Regular.ttf')
                if os.path.exists(font_path_regular):
                    font_size = int(63.7 * 1.33)
                    font = ImageFont.truetype(font_path_regular, font_size)
                else:
                    # Fallback fonts
                    try:
                        font = ImageFont.truetype("arialbd.ttf", 84)  # Arial Bold
                    except:
                        try:
                            font = ImageFont.truetype("arial.ttf", 84)
                        except:
                            try:
                                font = ImageFont.truetype("DejaVuSans-Bold.ttf", 84)
                            except:
                                try:
                                    font = ImageFont.truetype("DejaVuSans.ttf", 84)
                                except:
                                    # Final fallback to default
                                    font = ImageFont.load_default()
    except:
        font = ImageFont.load_default()

    date_text = event_date.strftime("%d %B %Y").upper()

    # Calculate text positioning (center within date coordinates)
    try:
        bbox = draw.textbbox((0, 0), date_text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
    except AttributeError:
        # Fallback for older PIL versions
        text_width = len(date_text) * 20  # Rough estimate
        text_height = 30

    # Center the text in the date coordinates
    date_x = date_coords[0] + (date_coords[2] - date_coords[0] - text_width) // 2
    date_y = date_coords[1] + (date_coords[3] - date_coords[1] - text_height) // 2

    # Draw the date text with stroke for extra thickness
    stroke_width = 3  # Thickness of the stroke outline
    
    # Draw stroke/outline first (black stroke around white text for better visibility)
    draw.text((date_x, date_y), date_text, fill="black", font=font, stroke_width=stroke_width, stroke_fill="black")
    
    # Draw main text on top
    draw.text((date_x, date_y), date_text, fill="white", font=font)

    # Add staff name text in the requested coordinates for day 1 and day 2
    # Use the same font family and size as the date text (variable `font` above)
    try:
        name_font = font
    except Exception:
        # Fallback
        try:
            name_font = ImageFont.truetype("arial.ttf", 48)
        except Exception:
            name_font = ImageFont.load_default()

    # Determine coordinates for name overlay (these are the requested coords)
    if staff.day_1 == 1:
        name_coords = (270, 291, 1339, 414)  # Updated for wider box
    else:
        name_coords = (270, 291, 1272, 416)  # Updated for wider box

    # Prepare name text and enforce 35-char limit
    staff_name_text = staff.name.strip() if staff.name else ''
    if len(staff_name_text) > 35:
        staff_name_text = staff_name_text[:35]

    max_width = name_coords[2] - name_coords[0]
    max_height = name_coords[3] - name_coords[1]

    # Start with the same font size as the date font (font variable above)
    base_size = getattr(font, 'size', None)
    if base_size is None:
        # fallback default
        base_size = 84

    # Try to locate a font path to recreate at different sizes
    font_path_candidate = None
    try:
        if hasattr(font, 'path'):
            font_path_candidate = font.path
    except Exception:
        font_path_candidate = None

    # fallback to Poppins-Regular if available
    if not font_path_candidate:
        candidate = os.path.join(settings.BASE_DIR, 'static/fonts/Poppins-Regular.ttf')
        if os.path.exists(candidate):
            font_path_candidate = candidate

    # If still not found, try Arial
    if not font_path_candidate:
        try:
            # System fonts may be found by name in truetype
            font_path_candidate = "arial.ttf"
        except Exception:
            font_path_candidate = None

    # Create a resizable font object; if not possible, fall back to existing name_font
    try:
        if font_path_candidate:
            test_font = ImageFont.truetype(font_path_candidate, base_size)
        else:
            test_font = name_font
    except Exception:
        test_font = name_font

    # Dynamically reduce font size until text fits within max_width and max_height
    chosen_font = test_font
    chosen_size = getattr(test_font, 'size', base_size)
    # Measure function
    def measure(text, fnt):
        try:
            b = draw.textbbox((0, 0), text, font=fnt)
            return b[2] - b[0], b[3] - b[1]
        except Exception:
            return draw.textsize(text, font=fnt)

    text_w, text_h = measure(staff_name_text, chosen_font)
    min_size = 12
    while (text_w > max_width or text_h > max_height) and chosen_size > min_size:
        chosen_size -= 2
        try:
            if font_path_candidate:
                chosen_font = ImageFont.truetype(font_path_candidate, chosen_size)
            else:
                # cannot resize default font, break
                break
        except Exception:
            break
        text_w, text_h = measure(staff_name_text, chosen_font)

    # If still doesn't fit and we couldn't resize further, fall back to truncation
    if (text_w > max_width or text_h > max_height) and chosen_size <= min_size:
        trimmed = staff_name_text
        while trimmed and measure(trimmed, chosen_font)[0] > max_width:
            trimmed = trimmed[:-1]
        staff_name_text = trimmed
        text_w, text_h = measure(staff_name_text, chosen_font)

    # Center the text in the name coordinates
    name_x = name_coords[0] + (max_width - text_w) // 2
    name_y = name_coords[1] + (max_height - text_h) // 2

    # Draw using chosen_font (which matches date font family if possible)
    try:
        draw.text((name_x, name_y), staff_name_text, font=chosen_font, fill='black', stroke_width=3, stroke_fill='black')
        draw.text((name_x, name_y), staff_name_text, font=chosen_font, fill='white')
    except TypeError:
        draw.text((name_x-1, name_y-1), staff_name_text, font=chosen_font, fill='black')
        draw.text((name_x, name_y), staff_name_text, font=chosen_font, fill='white')

    # Save the modified image to BytesIO
    output = BytesIO()
    template_img.save(output, format='PNG')
    output.seek(0)

    # Return the image as a download
    filename = f"staff_pass_{staff.name.replace(' ', '_').lower()}_{staff.department.replace(' ', '_').lower()}_day{staff.day_1}.png"
    response = HttpResponse(output.getvalue(), content_type='image/png')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.db import transaction
import pandas as pd
import random
import json

from .models import Staff, CheckIn, Winner, EventSettings
from .forms import StaffUploadForm, QRCodeScanForm, EventSettingsForm
from django.utils import timezone
from datetime import date

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid credentials. Please try again.')
    return render(request, 'lucky_draw/login.html')

@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

def require_event_dates(view_func):
    def _wrapped_view(request, *args, **kwargs):
        settings = EventSettings.get_solo()
        if not settings.day1_date or not settings.day2_date:
            return redirect('set_event_dates')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

@login_required
def set_event_dates(request):
    settings = EventSettings.get_solo()
    if request.method == 'POST':
        form = EventSettingsForm(request.POST, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, 'Event dates updated successfully.')
            return redirect('dashboard')
    else:
        form = EventSettingsForm(instance=settings)
    return render(request, 'lucky_draw/set_event_dates.html', {'form': form})

@login_required
def dashboard(request):
    settings = EventSettings.get_solo()
    if not settings.day1_date or not settings.day2_date:
        return redirect('set_event_dates')
    day1_checkins_count = CheckIn.objects.filter(day=1).count()
    day2_checkins_count = CheckIn.objects.filter(day=2).count()
    clear_db_env = os.getenv('CLEAR_DB_ENABLED', 'True').strip().lower()
    clear_db_enabled = clear_db_env in ['true', '1', 'yes', 'on']
    context = {
        'total_staff': Staff.objects.count(),
        'day1_checkins': day1_checkins_count,
        'day2_checkins': day2_checkins_count,
        'day1_winners': Winner.objects.filter(day=1).count(),
        'day2_winners': Winner.objects.filter(day=2).count(),
        'event_settings': settings,
        'can_draw_day1': day1_checkins_count >= 5,
        'can_draw_day2': day2_checkins_count >= 5,
        'draw_msg_day1': '' if day1_checkins_count >= 5 else 'At least 5 staff must be checked in for Day 1 to enable the draw.',
        'draw_msg_day2': '' if day2_checkins_count >= 5 else 'At least 5 staff must be checked in for Day 2 to enable the draw.',
        'clear_db_enabled': clear_db_enabled,
    }
    return render(request, 'lucky_draw/dashboard.html', context)

@login_required
def upload_staff(request):
    upload_form = StaffUploadForm()
    manual_form = StaffManualForm()
    if request.method == 'POST':
        if 'file' in request.FILES:
            upload_form = StaffUploadForm(request.POST, request.FILES)
            if upload_form.is_valid():
                try:
                    file = request.FILES['file']
                    # Read uploaded content into memory once (works for UploadedFile)
                    try:
                        file_bytes = file.read()
                    except Exception as e:
                        messages.error(request, f'Error reading uploaded file: {str(e)}')
                        return redirect('upload_staff')

                    # Try multiple encodings for CSV files to handle files exported from Windows/Excel
                    if file.name.lower().endswith('.csv'):
                        encodings = ['utf-8', 'utf-8-sig', 'cp1252', 'latin-1']
                        df = None
                        last_exc = None
                        for enc in encodings:
                            try:
                                decoded = file_bytes.decode(enc)
                                # Use StringIO so pandas can read from text
                                df = pd.read_csv(StringIO(decoded))
                                break
                            except Exception as e:
                                last_exc = e
                                continue
                        if df is None:
                            messages.error(request, f'Error processing file: unable to decode CSV. Please save the file as UTF-8 or try a different encoding. ({str(last_exc)})')
                            return redirect('upload_staff')
                    else:
                        # For Excel files, pass a BytesIO to pandas
                        try:
                            df = pd.read_excel(BytesIO(file_bytes))
                        except Exception as e:
                            messages.error(request, f'Error processing Excel file: {str(e)}')
                            return redirect('upload_staff')
                    import unicodedata
                    # Clear existing staff before import
                    Staff.objects.all().delete()
                    # Basic validation of dataframe shape
                    if df.shape[1] < 3:
                        messages.error(request, 'Uploaded file must have at least three columns: Staff Name, Department, Day')
                        return redirect('upload_staff')

                    def normalize_name(val):
                        if pd.isna(val):
                            return ''
                        s = str(val).strip()
                        # Normalize unicode to composed form and remove weird control characters
                        s = unicodedata.normalize('NFKC', s)
                        return s

                    for index, row in df.iterrows():
                        staff_name = normalize_name(row.iloc[0])
                        # Truncate imported names to 35 chars as required
                        if len(staff_name) > 35:
                            staff_name = staff_name[:35]
                        department = row.iloc[1]
                        day = int(row.iloc[2])
                        Staff.objects.create(
                            name=staff_name,
                            department=department,
                            day_1=day
                        )
                    messages.success(request, f'Successfully uploaded {len(df)} staff members')
                    return redirect('dashboard')
                except Exception as e:
                    messages.error(request, f'Error processing file: {str(e)}')
        elif 'manual_add' in request.POST:
            manual_form = StaffManualForm(request.POST)
            if manual_form.is_valid():
                manual_form.save()
                messages.success(request, 'Staff member added successfully.')
                return redirect('upload_staff')
    return render(request, 'lucky_draw/upload_staff.html', {'form': upload_form, 'manual_form': manual_form})

@login_required
@require_event_dates
def scan_qr(request):
    return render(request, 'lucky_draw/scan_qr.html')

@login_required
@csrf_exempt
@require_http_methods(["POST"])
@require_event_dates
def process_qr_scan(request):
    try:
        data = json.loads(request.body)
        qr_data = data.get('qr_data')
        
        if not qr_data:
            return JsonResponse({'success': False, 'message': 'No QR data provided'})
        
        staff_id, day = qr_data.split(':')
        staff = get_object_or_404(Staff, id=int(staff_id))
        day = int(day)
        settings = EventSettings.get_solo()
        today = date.today()
        # Validation: Only allow check-in for the correct day
        if day == 1 and today != settings.day1_date:
            return JsonResponse({'success': False, 'message': f'Check-in for Day 1 is only allowed on {settings.day1_date}.'})
        if day == 2 and today != settings.day2_date:
            return JsonResponse({'success': False, 'message': f'Check-in for Day 2 is only allowed on {settings.day2_date}.'})
        
        # Check if already checked in
        if CheckIn.objects.filter(staff=staff, day=day).exists():
            return JsonResponse({
                'success': False, 
                'message': f'{staff.name} has already been checked in for Day {day}'
            })
        
        # Create check-in
        CheckIn.objects.create(staff=staff, day=day)
        
        return JsonResponse({
            'success': True,
            'message': f'{staff.name} from {staff.department} successfully checked in for Day {day}',
            'staff_name': staff.name,
            'department': staff.department,
            'day': day
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def draw_winner(request):
    try:
        data = json.loads(request.body)
        day = int(data.get('day'))
        if day not in [1, 2]:
            return JsonResponse({'success': False, 'message': 'Invalid day'})
        checked_in_staff = list(CheckIn.objects.filter(day=day).values_list('staff_id', flat=True))
        existing_winners = list(Winner.objects.filter(day=day).values_list('staff_id', flat=True))
        # Enforce maximum of 5 winners per day
        if len(existing_winners) >= 5:
            return JsonResponse({'success': False, 'message': f'Maximum of 5 winners already drawn for Day {day}.'})
        if len(checked_in_staff) < 5:
            return JsonResponse({'success': False, 'message': f'At least 5 staff must be checked in for Day {day} to draw winners.'})
        available_staff = list(set(checked_in_staff) - set(existing_winners))
        if not available_staff:
            return JsonResponse({
                'success': False, 
                'message': f'No more available staff for Day {day} lucky draw'
            })
        # Use a lottery algorithm: shuffle and pick one
        import random
        random.shuffle(available_staff)
        winner_staff_id = available_staff[0]
        winner_staff = Staff.objects.get(id=winner_staff_id)
        next_order = Winner.objects.filter(day=day).count() + 1
        winner = Winner.objects.create(
            staff=winner_staff,
            day=day,
            draw_order=next_order
        )
        return JsonResponse({
            'success': True,
            'message': f'Winner #{next_order} for Day {day}: {winner_staff.name} from {winner_staff.department}',
            'winner_name': winner_staff.name,
            'department': winner_staff.department,
            'draw_order': next_order
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required
def view_checkins(request):
    day1_checkins = CheckIn.objects.filter(day=1).select_related('staff').order_by('-checked_in_at')
    day2_checkins = CheckIn.objects.filter(day=2).select_related('staff').order_by('-checked_in_at')

    context = {
        'day1_checkins': day1_checkins,
        'day2_checkins': day2_checkins,
    }
    return render(request, 'lucky_draw/view_checkins.html', context)

@login_required
def download_day1_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from django.utils import timezone

    # Create workbook
    wb = Workbook()

    # Sheet 1: Day 1 Checked-in staff
    ws_checked_in = wb.active
    ws_checked_in.title = "Day 1 - Checked In"

    # Sheet 2: Day 1 Not checked-in staff
    ws_not_checked_in = wb.create_sheet("Day 1 - Not Checked In")

    # Define headers
    checked_headers = ['Staff ID', 'Name', 'Department', 'Check-in Time']
    not_checked_headers = ['Staff ID', 'Name', 'Department']

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Function to style headers
    def style_headers(worksheet, headers, fill_color):
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = fill_color
            cell.alignment = Alignment(horizontal='center')

    # Function to auto-adjust column widths
    def adjust_column_widths(worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # Day 1 Checked In
    style_headers(ws_checked_in, checked_headers, header_fill)
    day1_checkins = CheckIn.objects.filter(day=1).select_related('staff').order_by('-checked_in_at')

    row_num = 2
    for checkin in day1_checkins:
        ws_checked_in.cell(row=row_num, column=1).value = checkin.staff.id
        ws_checked_in.cell(row=row_num, column=2).value = checkin.staff.name
        ws_checked_in.cell(row=row_num, column=3).value = checkin.staff.department
        ws_checked_in.cell(row=row_num, column=4).value = checkin.checked_in_at.strftime("%Y-%m-%d %H:%M:%S")
        row_num += 1
    adjust_column_widths(ws_checked_in)

    # Day 1 Not Checked In
    style_headers(ws_not_checked_in, not_checked_headers, header_fill)
    day1_staff = Staff.objects.filter(day_1=1)
    day1_checked_ids = CheckIn.objects.filter(day=1).values_list('staff_id', flat=True)

    row_num = 2
    for staff in day1_staff:
        if staff.id not in day1_checked_ids:
            ws_not_checked_in.cell(row=row_num, column=1).value = staff.id
            ws_not_checked_in.cell(row=row_num, column=2).value = staff.name
            ws_not_checked_in.cell(row=row_num, column=3).value = staff.department
            row_num += 1
    adjust_column_widths(ws_not_checked_in)

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    current_time = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"day1_checkin_report_{current_time}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)
    return response

@login_required
def download_day2_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from django.utils import timezone

    # Create workbook
    wb = Workbook()

    # Sheet 1: Day 2 Checked-in staff
    ws_checked_in = wb.active
    ws_checked_in.title = "Day 2 - Checked In"

    # Sheet 2: Day 2 Not checked-in staff
    ws_not_checked_in = wb.create_sheet("Day 2 - Not Checked In")

    # Define headers
    checked_headers = ['Staff ID', 'Name', 'Department', 'Check-in Time']
    not_checked_headers = ['Staff ID', 'Name', 'Department']

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")  # Green for Day 2

    # Function to style headers
    def style_headers(worksheet, headers, fill_color):
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = fill_color
            cell.alignment = Alignment(horizontal='center')

    # Function to auto-adjust column widths
    def adjust_column_widths(worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # Day 2 Checked In
    style_headers(ws_checked_in, checked_headers, header_fill)
    day2_checkins = CheckIn.objects.filter(day=2).select_related('staff').order_by('-checked_in_at')

    row_num = 2
    for checkin in day2_checkins:
        ws_checked_in.cell(row=row_num, column=1).value = checkin.staff.id
        ws_checked_in.cell(row=row_num, column=2).value = checkin.staff.name
        ws_checked_in.cell(row=row_num, column=3).value = checkin.staff.department
        ws_checked_in.cell(row=row_num, column=4).value = checkin.checked_in_at.strftime("%Y-%m-%d %H:%M:%S")
        row_num += 1
    adjust_column_widths(ws_checked_in)

    # Day 2 Not Checked In
    style_headers(ws_not_checked_in, not_checked_headers, header_fill)
    day2_staff = Staff.objects.filter(day_1=2)
    day2_checked_ids = CheckIn.objects.filter(day=2).values_list('staff_id', flat=True)

    row_num = 2
    for staff in day2_staff:
        if staff.id not in day2_checked_ids:
            ws_not_checked_in.cell(row=row_num, column=1).value = staff.id
            ws_not_checked_in.cell(row=row_num, column=2).value = staff.name
            ws_not_checked_in.cell(row=row_num, column=3).value = staff.department
            row_num += 1
    adjust_column_widths(ws_not_checked_in)

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    current_time = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"day2_checkin_report_{current_time}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)
    return response

@login_required
def view_winners(request):
    day1_winners = Winner.objects.filter(day=1).select_related('staff').order_by('draw_order')
    day2_winners = Winner.objects.filter(day=2).select_related('staff').order_by('draw_order')
    
    context = {
        'day1_winners': day1_winners,
        'day2_winners': day2_winners,
    }
    return render(request, 'lucky_draw/view_winners.html', context)

@login_required
def download_qr_codes(request):
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    staff_list = Staff.objects.all().order_by('name')
    total_staff_count = Staff.objects.count()
    departments = Staff.objects.values_list('department', flat=True).distinct()
    days = [1, 2]
    department_filter = request.GET.get('department', '')
    day_filter = request.GET.get('day', '')
    checked_in_filter = request.GET.get('checked_in', '')
    search_query = request.GET.get('search', '').strip()
    # Annotate staff with check-in status
    from django.db.models import Exists, OuterRef, Q
    staff_list = staff_list.annotate(
        checked_in=Exists(CheckIn.objects.filter(staff=OuterRef('pk'), day=OuterRef('day_1')))
    )
    if department_filter:
        staff_list = staff_list.filter(department=department_filter)
    if day_filter:
        staff_list = staff_list.filter(day_1=day_filter)
    if checked_in_filter == 'hide':
        staff_list = staff_list.filter(checked_in=False)
    # Enhanced search: prioritize exact matches for name, department, day, or ID
    if search_query:
        tokens = [t.strip().lower() for t in search_query.split() if t.strip()]
        queries = Q()
        for token in tokens:
            if token.isdigit():
                queries &= (Q(id=token) | Q(name__icontains=token) | Q(day_1=int(token)) | Q(department__icontains=token))
            else:
                queries &= (Q(name__icontains=token) | Q(department__icontains=token))
        staff_list = staff_list.filter(queries)

        # Annotate for exact match scoring
        from django.db.models import Case, When, IntegerField, Value
        exact_conditions = []
        for token in tokens:
            conds = []
            if token.isdigit():
                conds.append(When(id=token, then=Value(1)))
                conds.append(When(name__iexact=token, then=Value(1)))
                conds.append(When(day_1=int(token), then=Value(1)))
            else:
                conds.append(When(name__iexact=token, then=Value(1)))
                conds.append(When(department__iexact=token, then=Value(1)))
            exact_conditions.extend(conds)
        staff_list = staff_list.annotate(
            exact_match_score=Case(*exact_conditions, default=Value(0), output_field=IntegerField())
        )
        staff_list = staff_list.order_by('-exact_match_score', '-name', '-department', '-day_1')
    # Pagination - use dynamic items per page based on viewport
    page = request.GET.get('page', 1)
    
    # Get the per_page parameter from the request, with fallback to default
    try:
        per_page = int(request.GET.get('per_page', 12))
        # Constrain per_page to reasonable values to prevent abuse
        per_page = max(4, min(per_page, 48))  # Between 4 and 48 items
    except (ValueError, TypeError):
        # Default if not a valid number
        per_page = 12
    
    # Create paginator with dynamic page size
    paginator = Paginator(staff_list, per_page)
    
    try:
        staff_page = paginator.page(page)
    except PageNotAnInteger:
        staff_page = paginator.page(1)
    except EmptyPage:
        staff_page = paginator.page(paginator.num_pages)
    # Prepare staff list for JSON
    def staff_to_dict(staff):
        return {
            'id': staff.id,
            'name': staff.name,
            'department': staff.department,
            'day_1': staff.day_1,
            'checked_in': staff.checked_in,
        }
    staff_objs = list(staff_page.object_list) if staff_page else []
    staff_list_json = json.dumps([staff_to_dict(s) for s in staff_objs])
    if request.GET.get('ajax') == '1':
        return JsonResponse({
            'staff_list': [staff_to_dict(s) for s in staff_page.object_list],
            'page': staff_page.number,
            'num_pages': paginator.num_pages,
            'per_page': per_page,
            'total_count': paginator.count,
            'total_staff_count': total_staff_count,
            'departments': list(departments),
            'days': days,
        })
    return render(request, 'lucky_draw/download_qr_codes.html', {
        'staff_list': staff_page,
        'staff_list_json': staff_list_json,
        'departments': departments,
        'days': days,
        'department_filter': department_filter,
        'day_filter': day_filter,
        'checked_in_filter': checked_in_filter,
        'search_query': search_query,
        'paginator': paginator,
        'page_obj': staff_page,
        'total_staff_count': total_staff_count,
    })

@require_POST
@login_required
def delete_staff(request, staff_id):
    try:
        staff = Staff.objects.get(id=staff_id)
        staff.delete()
        return JsonResponse({'success': True})
    except Staff.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Staff not found.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def clear_database(request):
    clear_db_env = os.getenv('CLEAR_DB_ENABLED', 'True').strip().lower()
    clear_db_enabled = clear_db_env in ['true', '1', 'yes', 'on']
    if request.method == 'POST':
        if not clear_db_enabled:
            messages.error(request, 'Database clearing is disabled by admin.')
            return redirect('dashboard')
        Staff.objects.all().delete()
        CheckIn.objects.all().delete()
        Winner.objects.all().delete()
        EventSettings.objects.all().delete()
        messages.success(request, 'Database cleared!')
        return redirect('dashboard')
    return render(request, 'lucky_draw/clear_db_confirm.html', {'clear_db_enabled': clear_db_enabled})

@login_required
def download_template_image(request, day):
    try:
        from PIL import Image, ImageDraw, ImageFont
        import qrcode
        from io import BytesIO

        # Get event settings
        event_settings = EventSettings.get_solo()
        if not event_settings:
            return HttpResponse("Event settings not found. Please set event dates first.", status=400)

        # Get the date for the specific day
        if day == 1:
            event_date = event_settings.day1_date
            template_path = 'static/templates/day-1-lucky-draw.png'
            qr_coords = (504, 1569, 1083, 2151)
            date_coords = (459, 1355, 1125, 1446)
        elif day == 2:
            event_date = event_settings.day2_date
            template_path = 'static/templates/day-2-lucky-draw.png'
            qr_coords = (502, 1570, 1083, 2152)
            date_coords = (451, 1360, 1106, 1445)
        else:
            return HttpResponse("Invalid day specified.", status=400)

        if not event_date:
            return HttpResponse(f"Date for Day {day} not set.", status=400)

        # Load the template image
        template_full_path = os.path.join(settings.BASE_DIR, template_path)
        if not os.path.exists(template_full_path):
            return HttpResponse("Template image not found.", status=404)

        template_img = Image.open(template_full_path)
        draw = ImageDraw.Draw(template_img)

        # Generate QR code with event information
        qr_data = f"STAFF_LUCKY_DRAW_DAY_{day}_{event_date.strftime('%Y%m%d')}"
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")

        # Resize QR code to fit the coordinates
        qr_width = qr_coords[2] - qr_coords[0]
        qr_height = qr_coords[3] - qr_coords[1]
        qr_img = qr_img.resize((qr_width, qr_height), Image.Resampling.LANCZOS)

        # Paste QR code onto template
        template_img.paste(qr_img, (qr_coords[0], qr_coords[1]))

        # Add date text
        try:
            # Use Poppins font
            font_path = os.path.join(settings.BASE_DIR, 'static/fonts/Poppins-Regular.ttf')
            if os.path.exists(font_path):
                # Convert 63.7pt to pixels (approximately 84 pixels at 72 DPI)
                font_size = int(63.7 * 1.33)  # Rough conversion from pt to px
                font = ImageFont.truetype(font_path, font_size)
            else:
                # Fallback fonts
                try:
                    font = ImageFont.truetype("arial.ttf", 84)
                except:
                    try:
                        font = ImageFont.truetype("DejaVuSans.ttf", 84)
                    except:
                        # Final fallback to default
                        font = ImageFont.load_default()
        except:
            font = ImageFont.load_default()

        date_text = event_date.strftime("%d %B %Y").upper()
        bbox = draw.textbbox((0, 0), date_text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]

        # Center the text in the date coordinates
        date_x = date_coords[0] + (date_coords[2] - date_coords[0] - text_width) // 2
        date_y = date_coords[1] + (date_coords[3] - date_coords[1] - text_height) // 2

        # Draw the date text
        draw.text((date_x, date_y), date_text, fill="white", font=font)

        # Save the modified image to BytesIO
        output = BytesIO()
        template_img.save(output, format='PNG')
        output.seek(0)

        # Return the image as a download
        response = HttpResponse(output.getvalue(), content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename="day-{day}-lucky-draw-template.png"'
        return response

    except Exception as e:
        return HttpResponse(f"Error generating template image: {str(e)}", status=500)
